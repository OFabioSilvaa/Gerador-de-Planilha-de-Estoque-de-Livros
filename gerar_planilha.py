import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

NUM_LIVROS = 50
MARGEM_LUCRO = 0.30

#  Definindo Listas para geração de Dados Aleatórios
nomes_livros_base = [
    "A Arte da Guerra", "1984", "Dom Quixote", "O Pequeno Príncipe",
    "Cem Anos de Solidão", "Crime e Castigo", "Orgulho e Preconceito",
    "O Senhor dos Anéis", "Harry Potter e a Pedra Filosofal", "O Alquimista",
    "A Metamorfose", "Ulisses", "Moby Dick", "As Crônicas de Nárnia",
    "O Guia do Mochileiro das Galáxias", "O Apanhador no Campo de Centeio",
    "Em Busca do Tempo Perdido", "Grande Sertão: Veredas", "Memórias Póstumas de Brás Cubas",
    "Vidas Secas"
]

editoras = [
    "Editora Alfa", "Editora Beta", "Editora Gama", "Editora Delta",
    "HarperCollins", "Penguin Random House", "Rocco", "Companhia das Letras"
]
# Gerando dados aleatórios para um livro
def gerar_dados_livro():
    nome = random.choice(nomes_livros_base) + f" - Vol. {random.randint(1, 3)}" if random.random() < 0.3 else random.choice(nomes_livros_base)
    editora = random.choice(editoras)
    valor_compra = round(random.uniform(15.00, 150.00), 2)
    notas = random.randint(1, 5)
    quantidade = random.randint(1, 200)

    return {
        "Nome do Livro": nome,
        "Editora": editora,
        "Valor de Compra": valor_compra,
        "Notas": notas,
        "Quantidade": quantidade
    }

# Cria uma planilha Excel com dados de estoque de livros
def criar_planilha_estoque(nome_arquivo="estoque_livros.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque de Livros"

    headers = [
        "Nome do Livro", "Editora", "Valor de Compra", "Notas", "Quantidade",
        "Preço de Venda", "Lucro Total", "Valor Total"
    ]
    ws.append(headers)

    for _ in range(NUM_LIVROS):
        dados_livro = gerar_dados_livro()
        linha_dados = [
            dados_livro["Nome do Livro"],
            dados_livro["Editora"],
            dados_livro["Valor de Compra"],
            dados_livro["Notas"],
            dados_livro["Quantidade"]
        ]
        ws.append(linha_dados)
        current_row = ws.max_row

        # Fórmulas e Totais
        preco_venda_formula = f"=C{current_row}*(1+{MARGEM_LUCRO})"
        lucro_total_formula = f"=(F{current_row}-C{current_row})*E{current_row}"
        valor_total_formula = f"=F{current_row}*E{current_row}"

        ws[f"F{current_row}"] = preco_venda_formula
        ws[f"G{current_row}"] = lucro_total_formula
        ws[f"H{current_row}"] = valor_total_formula

    # Ajustando largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

# Salvando
    wb.save(nome_arquivo)
    print(f"Planilha '{nome_arquivo}' criada com sucesso!")

if __name__ == "__main__":
    criar_planilha_estoque()
